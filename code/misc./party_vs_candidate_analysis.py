"""
Party vs Candidate Specificity Analysis

This script is not part of the primary code for the paper but is included because
it is helpful for (1) seeing which races and districts had the greatest error, and
(2) comparing candidate-level to party-level error—large differences can be
indicative of data or benchmark issues.

Produces an Excel workbook with five sheets. Error/RMSE/Difference cells are
highlighted when the value is greater than 15 (or |Difference| > 15 where
applicable) to aid visualization.

  Sheet 1: RMSE by Year and Race
  Sheet 1b: House by District (candidate-level RMSE per district)
  Sheet 2: Top State Discrepancies
  Sheet 2b: Top House Districts (top 20 by RMSE per year)
  Sheet 3: Row-Level Comparison (party and candidate errors side by side; per-district for House)

Run from repo root or from code/misc/; paths are relative to tables_and_figures/.
"""

import numpy as np
import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def calculate_rmse(errors):
    """Calculate RMSE from a series of errors."""
    return np.sqrt(np.mean(errors**2))


def highlight_large_difference_column(ws, threshold=15):
    """Apply conditional formatting: highlight cells in 'Difference' column where |value| > threshold."""
    _highlight_column_by_condition(ws, 'Difference', f'ABS({{0}})>{threshold}')

def highlight_column_when_gt(ws, column_name, threshold=15):
    """Apply conditional formatting: highlight cells in column where value > threshold."""
    _highlight_column_by_condition(ws, column_name, f'{{0}}>{threshold}')

def _highlight_column_by_condition(ws, column_name, formula_template):
    """Find column by name in row 1 and apply conditional format; formula_template uses {0} for cell ref."""
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == column_name:
            col_letter = get_column_letter(col_idx)
            max_row = ws.max_row
            if max_row >= 2:
                fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
                formula = formula_template.format(f'{col_letter}2')
                ws.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}{max_row}',
                    FormulaRule(formula=[formula], fill=fill)
                )
            break


def main():
    from pathlib import Path
    script_dir = Path(__file__).resolve().parent
    # tables_and_figures is repo/tables_and_figures; repo is script's parent's parent when script is in code/misc
    repo_root = script_dir.parent.parent
    tables_and_figures_dir = repo_root / 'tables_and_figures'
    if not tables_and_figures_dir.exists():
        tables_and_figures_dir = Path.cwd() / 'tables_and_figures'
    # Output: same directory as this script (code/misc)
    out_dir = script_dir

    # ── Load and filter ──────────────────────────────────────────────────
    metrics_path = tables_and_figures_dir / 'input_tables' / 'metrics_base.xlsx'
    df = pd.read_excel(metrics_path)

    offices = ['President', 'U.S. Senate', 'U.S. House',
               'Governor', 'Attorney General', 'Secretary of State']

    df_cc = df[
        (df['Class'] == 'Candidate Choice') &
        (df['Weighting_Method'] == 'CES-Provided Weights') &
        (df['Validity_Scheme'] == 'All') &
        (df['Variable'].isin(offices)) &
        (df['Specificity'].isin(['Party', 'Candidate']))
    ].copy()

    print(f"Filtered data: {len(df_cc)} rows")
    print(f"Offices present: {sorted(df_cc['Variable'].unique())}")
    print(f"Years present: {sorted(df_cc['Year'].unique())}")

    # ── Sheet 1: RMSE by Year and Race ───────────────────────────────────
    sheet1_rows = []

    for office in offices:
        df_office = df_cc[df_cc['Variable'] == office]
        party_years = set(df_office[df_office['Specificity'] == 'Party']['Year'].unique())
        cand_years = set(df_office[df_office['Specificity'] == 'Candidate']['Year'].unique())
        overlap_years = party_years & cand_years

        for year in sorted(overlap_years):
            df_yr = df_office[df_office['Year'] == year]
            party_errors = df_yr[df_yr['Specificity'] == 'Party']['Error']
            cand_errors = df_yr[df_yr['Specificity'] == 'Candidate']['Error']

            rmse_party = calculate_rmse(party_errors)
            rmse_cand = calculate_rmse(cand_errors)

            sheet1_rows.append({
                'Year': year,
                'Office': office,
                'RMSE_Party': round(rmse_party, 4),
                'RMSE_Candidate': round(rmse_cand, 4),
                'Difference': round(rmse_cand - rmse_party, 4),
                'N_Observations_Party': len(party_errors),
                'N_Observations_Candidate': len(cand_errors),
            })

    sheet1 = pd.DataFrame(sheet1_rows)
    print(f"\nSheet 1 — RMSE by Year and Race: {len(sheet1)} rows")

    # ── Sheet 1b: U.S. House candidate-level RMSE by District (one row per district) ─────
    # No party comparison; just candidate-level error per district. Highlight RMSE > 15.
    house_district_rows = []
    df_house = df_cc[df_cc['Variable'] == 'U.S. House']
    if not df_house.empty and 'District' in df_house.columns:
        cand_house = df_house[(df_house['Specificity'] == 'Candidate') & (df_house['District'].notna())]
        if not cand_house.empty:
            for year in sorted(cand_house['Year'].unique()):
                cand_yr = cand_house[cand_house['Year'] == year]
                for state in cand_yr['State'].unique():
                    for district in cand_yr[cand_yr['State'] == state]['District'].unique():
                        cand_errors = cand_yr[(cand_yr['State'] == state) & (cand_yr['District'] == district)]['Error']
                        if len(cand_errors) < 1:
                            continue
                        house_district_rows.append({
                            'Year': year,
                            'State': state,
                            'District': district,
                            'RMSE': round(calculate_rmse(cand_errors), 4),
                            'N_Observations': len(cand_errors),
                        })
    sheet1b = pd.DataFrame(house_district_rows)
    if not sheet1b.empty:
        sheet1b = sheet1b.sort_values(['Year', 'RMSE'], ascending=[True, False])
    print(f"Sheet 1b — House by District: {len(sheet1b)} rows")

    # ── Sheet 2: Top State Discrepancies ─────────────────────────────────
    sheet2_rows = []

    for office in offices:
        df_office = df_cc[df_cc['Variable'] == office]
        party_years = set(df_office[df_office['Specificity'] == 'Party']['Year'].unique())
        cand_years = set(df_office[df_office['Specificity'] == 'Candidate']['Year'].unique())
        overlap_years = party_years & cand_years

        for year in sorted(overlap_years):
            df_yr = df_office[df_office['Year'] == year]

            # Get states with data for both specificities
            party_states = set(df_yr[df_yr['Specificity'] == 'Party']['State'].unique())
            cand_states = set(df_yr[df_yr['Specificity'] == 'Candidate']['State'].unique())
            common_states = party_states & cand_states

            state_discreps = []
            for state in common_states:
                df_state = df_yr[df_yr['State'] == state]
                party_errors = df_state[df_state['Specificity'] == 'Party']['Error']
                cand_errors = df_state[df_state['Specificity'] == 'Candidate']['Error']

                # Require at least 1 observation per specificity
                if len(party_errors) < 1 or len(cand_errors) < 1:
                    continue

                rmse_p = calculate_rmse(party_errors)
                rmse_c = calculate_rmse(cand_errors)

                state_discreps.append({
                    'Year': year,
                    'Office': office,
                    'State': state,
                    'RMSE_Party': round(rmse_p, 4),
                    'RMSE_Candidate': round(rmse_c, 4),
                    'Difference': round(rmse_c - rmse_p, 4),
                    'N_Party': len(party_errors),
                    'N_Candidate': len(cand_errors),
                    '_abs_diff': abs(rmse_c - rmse_p),
                })

            # Rank by |discrepancy|, take top 10
            state_discreps.sort(key=lambda x: x['_abs_diff'], reverse=True)
            for rank, row in enumerate(state_discreps[:10], 1):
                row['Rank'] = rank
                sheet2_rows.append(row)

    sheet2 = pd.DataFrame(sheet2_rows)
    if not sheet2.empty:
        sheet2 = sheet2[['Year', 'Office', 'Rank', 'State',
                         'RMSE_Party', 'RMSE_Candidate', 'Difference',
                         'N_Party', 'N_Candidate']]

    print(f"Sheet 2 — Top State Discrepancies: {len(sheet2)} rows")

    # ── Sheet 2b: Top U.S. House districts by candidate RMSE (top 20 per year) ─────────
    sheet2b_rows = []
    if not sheet1b.empty:
        for year in sheet1b['Year'].unique():
            df_yr = sheet1b[sheet1b['Year'] == year].nlargest(20, 'RMSE')
            for rank, (_, row) in enumerate(df_yr.iterrows(), 1):
                sheet2b_rows.append({
                    'Year': year,
                    'Rank': rank,
                    'State': row['State'],
                    'District': row['District'],
                    'RMSE': row['RMSE'],
                    'N_Observations': row['N_Observations'],
                })
    sheet2b = pd.DataFrame(sheet2b_rows)
    print(f"Sheet 2b — Top House Districts: {len(sheet2b)} rows")

    # ── Sheet 3: Row-Level Comparison ────────────────────────────────────
    # Merge party and candidate rows on shared keys so errors appear
    # side by side. For U.S. House, merge by district so one row per district.

    merge_cols = ['Year', 'State', 'Variable', 'Category']
    # For House we need District in the key so we get one row per district
    party_for_merge = df_cc[df_cc['Specificity'] == 'Party'].copy()
    cand_for_merge = df_cc[df_cc['Specificity'] == 'Candidate'].copy()
    party_for_merge['_District'] = party_for_merge.apply(
        lambda r: r['District'] if r['Variable'] == 'U.S. House' else '',
        axis=1
    )
    cand_for_merge['_District'] = cand_for_merge.apply(
        lambda r: r['District'] if r['Variable'] == 'U.S. House' else '',
        axis=1
    )
    merge_cols_with_dist = merge_cols + ['_District']

    party_rows = party_for_merge[
        merge_cols_with_dist + ['District', 'Benchmark', 'CES_Weighted', 'Error', 'n_respondents']
    ].rename(columns={
        'District': 'District_Party',
        'Benchmark': 'Benchmark_Party',
        'CES_Weighted': 'CES_Party',
        'Error': 'Error_Party',
        'n_respondents': 'N_Party',
    })

    cand_rows = cand_for_merge[
        merge_cols_with_dist + ['District', 'Benchmark', 'CES_Weighted', 'Error', 'n_respondents']
    ].rename(columns={
        'District': 'District_Candidate',
        'Benchmark': 'Benchmark_Candidate',
        'CES_Weighted': 'CES_Candidate',
        'Error': 'Error_Candidate',
        'n_respondents': 'N_Candidate',
    })

    sheet3 = party_rows.merge(cand_rows, on=merge_cols_with_dist, how='outer')
    sheet3 = sheet3.drop(columns=['_District'])

    # Flag whether the errors differ
    sheet3['Errors_Match'] = np.isclose(
        sheet3['Error_Party'].fillna(float('inf')),
        sheet3['Error_Candidate'].fillna(float('inf')),
        atol=1e-6,
    )

    # Sort for readability (by district for House)
    office_order = {o: i for i, o in enumerate(offices)}
    sheet3['_office_sort'] = sheet3['Variable'].map(office_order)
    sheet3 = sheet3.sort_values(
        ['Year', '_office_sort', 'State', 'Category', 'District_Party']
    ).drop(columns='_office_sort')

    # Rename Variable -> Office for consistency with other sheets
    sheet3 = sheet3.rename(columns={'Variable': 'Office'})

    print(f"Sheet 3 — Row-Level Comparison: {len(sheet3)} rows "
          f"({(~sheet3['Errors_Match']).sum()} with differing errors)")

    # ── Write Excel ──────────────────────────────────────────────────────
    out_path = out_dir / 'party_vs_candidate_by_year_state.xlsx'
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        sheet1.to_excel(writer, sheet_name='RMSE by Year and Race', index=False)
        if not sheet1b.empty:
            sheet1b.to_excel(writer, sheet_name='House by District', index=False)
        sheet2.to_excel(writer, sheet_name='Top State Discrepancies', index=False)
        if not sheet2b.empty:
            sheet2b.to_excel(writer, sheet_name='Top House Districts', index=False)
        sheet3.to_excel(writer, sheet_name='Row-Level Comparison', index=False)
        # Highlight error/RMSE/Difference when > 15 (or |Difference| > 15) on all 5 sheets
        th = 15
        if 'RMSE by Year and Race' in writer.book.sheetnames:
            highlight_large_difference_column(writer.book['RMSE by Year and Race'], threshold=th)
        if 'House by District' in writer.book.sheetnames:
            highlight_column_when_gt(writer.book['House by District'], 'RMSE', threshold=th)
        if 'Top State Discrepancies' in writer.book.sheetnames:
            highlight_large_difference_column(writer.book['Top State Discrepancies'], threshold=th)
        if 'Top House Districts' in writer.book.sheetnames:
            highlight_column_when_gt(writer.book['Top House Districts'], 'RMSE', threshold=th)
        if 'Row-Level Comparison' in writer.book.sheetnames:
            highlight_column_when_gt(writer.book['Row-Level Comparison'], 'Error_Party', threshold=th)
            highlight_column_when_gt(writer.book['Row-Level Comparison'], 'Error_Candidate', threshold=th)

    print(f"\nSaved: {out_path}")

    # ── Quick summary ────────────────────────────────────────────────────
    print("\n--- Year/Office RMSE Summary ---")
    if not sheet1.empty:
        for _, r in sheet1.iterrows():
            print(f"  {r['Year']}  {r['Office']:<20s}  "
                  f"Party={r['RMSE_Party']:6.2f}  Cand={r['RMSE_Candidate']:6.2f}  "
                  f"Diff={r['Difference']:+6.2f}")


if __name__ == '__main__':
    main()
